# -*- coding: utf-8 -*-
import importlib,sys 
import time
import urllib
import requests
import numpy as np
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl import Workbook

importlib.reload(sys)

#Some User Agents
hds=[{'User-Agent': 'Mozilla/5.0 (Windows NT 6.2; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0'},\
{'User-Agent': 'Mozilla/5.0 (Android; Mobile; rv:14.0) Gecko/14.0 Firefox/14.0'}
]

def book_spider(book_tag):
    page_num=0;
    book_list=[]
    try_times=0
    
    while(1):
        url='http://www.douban.com/tag/'+urllib.request.quote(book_tag)+'/book?start='+str(page_num*15)
        time.sleep(np.random.rand()*5)
        
        #Last Version
        try:
            req = urllib.request.Request(url, headers=hds[page_num%len(hds)])
            plain_text = urllib.request.urlopen(req).read()  
        except:
            print ("Error Request().")
            continue
  
        # #Previous Version, IP is easy to be Forbidden
        # source_code = requests.get(url) 
        # plain_text = source_code.text  
        
        soup = BeautifulSoup(plain_text, from_encoding='utf-8')
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip() #标题
            desc = book_info.find('div', {'class':'desc'}).string.strip() #作者/译者 和 出版信息
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href')
            
            print(title)

            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
                if int(people_num) < 1000:
                    continue
            except:
                people_num ='0'
            try:
                author_info = '/'.join(desc_list[0:-3])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print ('Downloading Information From Page %d' % (page_num))
    return book_list

def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        plain_text = urllib.request.urlopen(req).read()
    except:
        print("Error get_people_num().")
    soup = BeautifulSoup(plain_text)
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num

def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)
    return book_lists

def print_book_lists_excel(book_lists,book_tag_lists):
    wb=Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i])) #utf8->unicode
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)

# def output_html(book_lists):
#         count=1
#         fout = open('output.html', 'w',encoding='utf-8')
#         fout.write('<html>')
#         fout.write('<body>')
#         fout.write('<table>')
#         #ascii
#         for bl in book_lists[0]:
#             fout.write('<tr>')
#             fout.write("<td>%d</td>"%count)
#             fout.write("<td>%s</td>"%bl[0])
#             fout.write("<td>%f</td>"%float(bl[1]))
#             fout.write("<td>%s</td>"%bl[2])
#             fout.write("<td>%s</td>"%bl[3])
#             fout.write('</tr>')
#             count+=1
#         fout.write('</table>')
#         fout.write('</body>')
#         fout.write('</html>')
#         fout.close()

if __name__=='__main__':
    book_tag_lists = ['小说','计算机','历史']
    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)


